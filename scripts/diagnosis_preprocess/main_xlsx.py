import argparse
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from time import sleep

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook
from preprocess_prompt import (
    ADDITIONAL_INFO_PROMPT,
    AUXILIARY_EXAMINATION_PROMPT,
    PERSONAL_HISTORY_PROMPT,
    SYMPTOM_PROMPT,
)
from tqdm.asyncio import tqdm

load_dotenv()

# client = OpenAI(
#     api_key=os.getenv("OPENAI_API_KEY"),
#     base_url=os.getenv("OPENAI_API_BASE_URL"),
# )
# model_name = "gpt-4o"

client = OpenAI(
    api_key="0",
    base_url="http://localhost:8000/v1",
)
model_name = "Qwen/Qwen3-8B"

def get_llm_output(prompt: str) -> dict:
    cnt = 0
    while True:
        cnt += 1
        try:
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {
                        "role": "user",
                        "content": prompt,
                    }
                ],
            )
            content = response.choices[0].message.content
            if "</think>" in content:
                content = content.split("</think>")[1].strip()
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            # print(content)
            return json.loads(content)
        except Exception as e:
            print(f"Error during LLM call: {e}")
            if cnt >= 5:
                return {}
            print(f"Retrying LLM call, sleep time: {2 ** cnt - 1} seconds")
            sleep(2 ** cnt - 1)

def pick(label: str, next_labels: list[str], *, text: str) -> str:
    """
    抽取形如：<label>：<内容>，直到遇到 next_labels 里的任意一个标签或文末。
    """
    # 允许中文/英文冒号，允许 label 前有空白或制表符
    # 内容用非贪婪匹配，遇到下一个标签就停止
    next_alt = "|".join(map(re.escape, next_labels)) if next_labels else r"$"
    pattern = rf"{re.escape(label)}\s*[：:\n]\s*(.*?)(?=\n\s*(?:{next_alt})\s*[：:\n]|\Z)"
    m = re.search(pattern, text, flags=re.S)
    if not m:
        return ""
    # 清理：去掉多余空白行、行尾空格
    s = m.group(1)
    s = re.sub(r"[ \t]+\n", "\n", s)         # 行尾空格
    s = re.sub(r"\n{3,}", "\n\n", s)         # 过多空行
    return s.strip()

def extract_sections(text: str) -> dict:

    name = pick("姓名", ["职业", "性别", "年龄"], text=text)
    age = pick("年龄", ["入院时间", "联系电话", "性别"], text=text)
    gender = pick("性别", ["联系电话", "年龄", "现住址"], text=text)

    symptom_text = pick("主诉", ["既往史"], text=text)
    personal_history_text = pick("个人史", ["婚育史", "家族史", "体格检查", "体 格 检 查", "体  格  检  查"], text=text)
    marriage_childbearing_history = pick("婚育史", ["家族史", "体格检查", "体 格 检 查", "体  格  检  查"], text=text)
    family_history = pick("家族史", ["体格检查", "体 格 检 查", "体  格  检  查"], text=text)

    if "体格检查" in text:
        text = text.replace("体格检查", "体  格  检  查")
    if "体 格 检 查" in text:
        text = text.replace("体 格 检 查", "体  格  检  查")
    basic_info = pick("体  格  检  查", ["一般情况"], text=text)

    general_condition = pick("一般情况", ["专科情况", "专 科 检 查 ", "辅助检查", "辅 助 检 查", "辅  助  检  查"], text=text)

    if "专 科 检 查" in text:
        text = text.replace("专 科 检 查", "专科情况")
    special_examination = pick("专科情况", ["辅助检查", "辅 助 检 查", "辅  助  检  查"], text=text)

    if "辅助检查" in text:
        text = text.replace("辅助检查", "辅  助  检  查")
    if "辅 助 检 查" in text:
        text = text.replace("辅 助 检 查", "辅  助  检  查")
    auxiliary_examination_text = pick("辅  助  检  查", ["初步诊断"], text=text)

    with ThreadPoolExecutor(max_workers=3) as executor:
        future_symptom = executor.submit(
            get_llm_output,
            SYMPTOM_PROMPT.format(symptom_text=symptom_text),
        )
        future_personal_history = executor.submit(
            get_llm_output,
            PERSONAL_HISTORY_PROMPT.format(personal_history_text=personal_history_text),
        )
        future_auxiliary_examination = executor.submit(
            get_llm_output,
            AUXILIARY_EXAMINATION_PROMPT.format(auxiliary_examination_text=auxiliary_examination_text),
        )

        symptom = future_symptom.result()
        personal_history = future_personal_history.result()
        auxiliary_examination = future_auxiliary_examination.result()
    
    return {
        "personal_info": {
            "demographics": {
                "name": name,
                "gender": gender,
                "age": age,
            },
            "personal_history": {
                "smoking_status": personal_history.get("smoking_status", "未知"),
                "alcohol_use": personal_history.get("alcohol_use", "未知"),
                "family_history": family_history,
                "marriage_childbearing_history": marriage_childbearing_history
            }
        },
        "symptom": {
            "symptom_duration": symptom.get("symptom_duration", ""),
            "chief_complaint": symptom.get("chief_complaint", ""),
            "additional_symptom": symptom.get("additional_symptom", "")
        },
        "physical_examination": {
            "basic_info": basic_info,
            "general_condition": general_condition,
            "special_examination": special_examination,
        },
        **auxiliary_examination,
        "diagnosis": {},
        "treatment": {},
    }



def add_extra_info(text: str, existing_json: dict) -> dict:
    def _f(text: str):
        pattern = r"患者(?P<name>[^，,]+)[，,]\s*(?P<gender>男|女)[，,]\s*(?P<age>\d+)岁"
        
        match = re.search(pattern, text)
        if not match:
            return None, None, None
        
        name = match.group("name")
        gender = match.group("gender")
        age = int(match.group("age"))
        
        return name, gender, age
    name, gender, age = _f(text)

    results = get_llm_output(
        ADDITIONAL_INFO_PROMPT.format(
            text=text,
            symptom=json.dumps(existing_json.get("symptom", {})),
            auxiliary_examination=json.dumps(existing_json.get("auxiliary_examination", [])),
        )
    )
    current_json = {
        "symptom": {
            "chief_complaint": results.get("chief_complaint", ""),
            "additional_symptom": results.get("additional_symptom", ""),
            "symptom_duration": results.get("symptom_duration", ""),
        },
        "diagnosis": results.get("diagnosis", {}),
        "treatment": results.get("treatment", {}),
        "auxiliary_examination": results.get("auxiliary_examination", []),
    }
    # print(f"Diagnosis extracted: {current_json['diagnosis']}")
    existing_json.update(current_json)
    existing_json["personal_info"]["demographics"] = {
        "name": name,
        "gender": gender,
        "age": f"{age}岁",
    }
    return existing_json


def process_patient(patient_id: int, chunks: list[str]):
    processed_json = {}

    flag1, flag2 = False, False

    for chunk in chunks:
        if "体 格 检 查" in chunk or "体  格  检  查" in chunk:
            processed_json = extract_sections(chunk)
            flag1 = True
        elif "一、入院情况" in chunk:
            processed_json = add_extra_info(chunk, processed_json)
            flag2 = True
        if flag1 and flag2:
            break
    if (not flag1) or (not flag2):
        return patient_id, processed_json, ValueError(f"Could not find expected sections in patient {patient_id}")

    return patient_id, processed_json, None

def main():
    parser = argparse.ArgumentParser(description="Preprocess patient diagnosis data.")
    parser.add_argument("--input_file", type=str, required=True)
    parser.add_argument("--output_dir", type=str, required=True)
    parser.add_argument("--workers", type=int, default=16)
    args = parser.parse_args()

    input_file = Path(args.input_file)
    output_dir = Path(args.output_dir)

    output_success_dir = output_dir / "success"
    output_error_dir = output_dir / "error"
    output_success_dir.mkdir(parents=True, exist_ok=True)
    output_error_dir.mkdir(parents=True, exist_ok=True)
    
    wb = load_workbook(input_file)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    # rows 能按照 rows['patientid'] 为第一关键字，rows['病历时间'] 作为第二关键字排序吗？
    sorted_rows = sorted(rows, key=lambda row: (dict(zip(headers, row, strict=False)).get("patientid", ""), dict(zip(headers, row, strict=False)).get("病历时间", "")))

    all_data = {}
    cnt = 0
    for row in sorted_rows:
        cnt += 1
        row_dict = dict(zip(headers, row, strict=False))
        patient_id = row_dict.get("patientid")
        input_text = row_dict.get("文书内容", "")
        # print(f"Processing patient {patient_id}: {input_text}")
        if patient_id in all_data:
            all_data[patient_id].append(input_text)
        else:
            all_data[patient_id] = [input_text]
    
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(process_patient, patient_id, all_data[patient_id]): patient_id
            for patient_id in all_data.keys()
        }

        for future in tqdm(as_completed(futures), total=len(futures), desc="Processing patients"):
            patient_id = futures[future]
            try:
                patient_id, processed_json, error = future.result()
                if error is not None:
                    raise RuntimeError(f"Error processing patient {patient_id}: {error}")
                if processed_json.get("symptom", {}).get("chief_complaint", "") == "":
                    raise RuntimeError(f"Chief complaint is empty for patient {patient_id}")
                if processed_json.get("diagnosis", {}).get("disease_name", "") == "":
                    raise RuntimeError(f"Disease name is empty for patient {patient_id}")
                if processed_json.get("diagnosis", {}).get("stage", "") == "":
                    raise RuntimeError(f"Stage is empty for patient {patient_id}")
                if processed_json.get("treatment", {}).get("plan", "") == "":
                    raise RuntimeError(f"Treatment plan is empty for patient {patient_id}")
                
                if processed_json.get("personal_info", {}).get("personal_history", {}).get("smoking_status", "未知") == "未知":
                    raise RuntimeError(f"Smoking status is unknown for patient {patient_id}")
                if processed_json.get("personal_info", {}).get("personal_history", {}).get("alcohol_use", "未知") == "未知":
                    raise RuntimeError(f"Alcohol use is unknown for patient {patient_id}")
                
                with (output_success_dir / f"{patient_id}.json").open("w", encoding="utf-8") as f:
                    json.dump(processed_json, f, ensure_ascii=False, indent=4)
            except Exception as e:
                with (output_error_dir / f"{patient_id}_error.txt").open("w", encoding="utf-8") as f:
                    f.write(str(e))


if __name__ == "__main__":
    main()
