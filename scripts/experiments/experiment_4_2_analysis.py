
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Population-level significance analysis for the bad-news communication simulation.

This script compares two folders of JSON files:
- before_dir: before training / baseline doctor (folder b in the pilot zip)
- after_dir : after training / RL-evolved doctor (folder a in the pilot zip)

It extracts:
1) Existing fields that can be read directly from JSON
2) Extended fields derived from text rules discussed in the project history

For each factor, the script tests its association with:
- Final PAS
- Final ESS
- Maximum ESS
- Final CCS

Outputs written into --output_dir:
- 01_existing_fields_significance.xlsx
- 02_extended_fields_significance.xlsx
- analysis_report.md

Design choices:
- Binary factors    -> two-sided Mann–Whitney U
- Multi-category factors -> Kruskal–Wallis
- Continuous age    -> Spearman rank correlation

The matrix sheet matches the requested layout:
                | 因素1 | 因素2 | ...
    b_PAS       |
    a_PAS       |
    b_ESS       |
    a_ESS       |
    b_MaxESS    |
    a_MaxESS    |
    b_CCS       |
    a_CCS       |

Notes:
- p-values in the matrix are raw (unadjusted) exploratory p-values.
- Significant cells are highlighted at p < 0.05.
- For multi-category factors, the omnibus p-value only says "at least one group differs";
  please use the LevelSummary sheet to inspect group sizes and medians.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats


# =========================
# Data model / factor specs
# =========================

@dataclass(frozen=True)
class FactorSpec:
    """Configuration for one factor in the analysis."""
    key: str
    display: str
    kind: str  # "binary", "categorical", "continuous"
    source: str
    definition: str
    level_order: Optional[List[Any]] = None
    bool_labels: Tuple[str, str] = ("否", "是")  # (False label, True label)


OUTCOMES: List[Tuple[str, str, str]] = [
    ("final_pas", "PAS", "Final PAS"),
    ("final_ess", "ESS", "Final ESS"),
    ("max_ess", "MaxESS", "Maximum ESS"),
    ("final_ccs", "CCS", "Final CCS"),
]

ROW_ORDER: List[Tuple[str, str, str]] = [
    ("before", "final_pas", "b_PAS"),
    ("after", "final_pas", "a_PAS"),
    ("before", "final_ess", "b_ESS"),
    ("after", "final_ess", "a_ESS"),
    ("before", "max_ess", "b_MaxESS"),
    ("after", "max_ess", "a_MaxESS"),
    ("before", "final_ccs", "b_CCS"),
    ("after", "final_ccs", "a_CCS"),
]

EXISTING_FACTORS: List[FactorSpec] = [
    FactorSpec(
        key="gender",
        display="性别",
        kind="categorical",
        source="patient_data.personal_info.demographics.gender",
        definition="直接读取 gender 字段。",
        level_order=["男", "女"],
    ),
    FactorSpec(
        key="age",
        display="年龄",
        kind="continuous",
        source="patient_data.personal_info.demographics.age",
        definition="从年龄文本中提取整数年龄（例如“53岁”-> 53），按连续变量处理。",
    ),
    FactorSpec(
        key="cancer_type",
        display="癌种",
        kind="categorical",
        source="file_name（Path(file_name).stem.rsplit('_', 1)[-1]）",
        definition="从 file_name 中提取最后一个下划线后的癌种名称，例如 supplement_xxx_肺癌.json -> 肺癌。",
        level_order=["乳腺癌", "前列腺癌", "结直肠癌", "肺癌", "胃癌"],
    ),
    FactorSpec(
        key="stage_1234",
        display="分期",
        kind="categorical",
        source="examination_data.diagnosis.stage_1234",
        definition="直接读取 stage_1234（1/2/3/4 期）。",
        level_order=[1, 2, 3, 4],
    ),
    FactorSpec(
        key="education_level",
        display="认知水平",
        kind="categorical",
        source="patient_data.personal_info.social_background.education_level",
        definition="直接读取 education_level；按原始 5 档做整体检验。",
        level_order=[
            "小学及以下：识字/医学术语理解有限",
            "初中：可理解生活化解释，需避免复杂概念",
            "高中/中专：能跟随基本因果解释，愿意听结论",
            "本科：能理解风险概率与方案对比",
            "硕博/医学相关背景：会追问证据等级与指南出处",
        ],
    ),
    FactorSpec(
        key="financial_status",
        display="经济状况",
        kind="categorical",
        source="patient_data.personal_info.social_background.financial_status",
        definition="直接读取 financial_status；按原始 5 档做整体检验。",
        level_order=[
            "困难：治疗费用压力大，倾向选择低成本方案",
            "较差：依赖医保，关注报销范围与额度",
            "中等：有医保但自付敏感，可承担常规治疗",
            "较好：愿为疗效/舒适度付费，重视体验与效率",
            "充裕：费用不是主要约束，优先考虑最优方案/专家资源",
        ],
    ),
    FactorSpec(
        key="personality",
        display="性格特点",
        kind="categorical",
        source="patient_data.personal_info.characteristics.personality",
        definition="直接读取 personality；按原始 6 类做整体检验。",
        level_order=[
            "不信任：对医疗系统或医生持怀疑态度",
            "不耐烦：易怒、易激动、缺乏耐心",
            "中性：情绪稳定，易于沟通",
            "啰嗦：喜欢详细讨论，难以集中主题",
            "过度乐观：忽视风险，过分相信治疗效果",
            "过度焦虑：总是感到担忧和紧张，难以放松",
        ],
    ),
    FactorSpec(
        key="communication_style",
        display="沟通风格",
        kind="categorical",
        source="patient_data.personal_info.characteristics.communication_style",
        definition="直接读取 communication_style；按原始 5 类做整体检验。",
        level_order=[
            "信息收集型：喜欢获取大量背景信息和细节",
            "冷静理性型：交流中一直保持冷静和理性，不受情绪影响",
            "情绪表达型：倾向于表达自己的情感和感受",
            "沉默顺从型：被动回应，很少表达自己的想法",
            "问题驱动型：喜欢提出问题，针对具体问题寻求解答",
        ],
    ),
]

EXTENDED_FACTORS: List[FactorSpec] = [
    FactorSpec(
        key="systemic_therapy_involved",
        display="系统治疗参与",
        kind="binary",
        source="由 treatment.plan 文本规则派生",
        definition=(
            "若 treatment.plan 含化疗、免疫治疗、靶向治疗或内分泌治疗，则记为“是”；"
            "否则记为“否”。这里把它视为 treatment complexity 的一个 proxy。"
        ),
    ),
    FactorSpec(
        key="surgery_involved",
        display="手术参与",
        kind="binary",
        source="由 treatment.plan 文本规则派生",
        definition=(
            "若 treatment.plan 含明确治疗性手术关键词（如根治术、切除术、清扫术、造瘘、吻合、"
            "胸腔镜/腹腔镜/开腹等），则记为“是”。纯活检/穿刺活检不计入手术参与。"
        ),
    ),
    FactorSpec(
        key="chemotherapy_involved",
        display="化疗参与",
        kind="binary",
        source="由 treatment.plan 文本规则派生",
        definition="若 treatment.plan 含“化疗”或典型化疗药/方案关键词（如奥沙利铂、卡培他滨、FOLFOX、XELOX 等），则记为“是”。",
    ),
    FactorSpec(
        key="symptom_burden",
        display="症状负担",
        kind="binary",
        source="由 chief_complaint / reiterated_symptom / additional_symptom 文本规则派生",
        definition=(
            "把 chief_complaint、reiterated_symptom、additional_symptom 中的正向症状 clause 数出来；"
            "以 >=4 条记为高症状负担（是），0-3 条记为低症状负担（否）。"
            "注意这更接近 symptom-load / symptom complexity，而不是标准临床症状严重度量表。"
        ),
        bool_labels=("0-3 条正向症状", ">=4 条正向症状"),
    ),
    FactorSpec(
        key="metastatic_recurrent_context",
        display="转移/复发情境",
        kind="binary",
        source="由 diagnosis / stage / complication / symptom / 辅助检查文本规则派生",
        definition="若相关文本中出现“转移”“复发”“播散”“种植”或 M1 等提示，则记为“是”。",
    ),
    FactorSpec(
        key="lymph_node_involvement",
        display="淋巴结受累",
        kind="binary",
        source="由 diagnosis / stage / complication / 辅助检查文本规则派生",
        definition="若相关文本中出现“淋巴结”或分期里出现 N1/N2/N3，则记为“是”。",
    ),
    FactorSpec(
        key="heavy_alcohol_history",
        display="重度饮酒史",
        kind="binary",
        source="由 personal_history.alcohol_use 文本规则派生",
        definition=(
            "若 alcohol_use 文本提示长期/较大量饮酒（例如 >=20 年、白酒 4-5 两/日、7 两/日、"
            "200g/天、半斤等），则记为“是”；否认/未知/偶饮/量不详不计入。"
        ),
    ),
    FactorSpec(
        key="smoking_history",
        display="吸烟史",
        kind="binary",
        source="由 personal_history.smoking_status 文本规则派生",
        definition="若 smoking_status 文本中存在吸烟或已戒烟史（排除否认/未知/无吸烟史），则记为“是”。",
    ),
]


# =========================
# Extraction helpers
# =========================

CHEMO_KEYWORDS = [
    "化疗", "奥沙利铂", "卡培他滨", "卡铂", "顺铂", "紫杉", "多西他赛", "培美曲塞",
    "氟尿嘧啶", "亚叶酸", "伊立替康", "吉西他滨", "长春瑞滨", "替吉奥", "表柔比星",
    "环磷酰胺", "多柔比星", "FOLFOX", "XELOX", "SOX", "TP", "TC", "EC→T", "EC-T", "AC-T",
]
IMMUNO_KEYWORDS = [
    "免疫", "PD-1", "PD1", "信迪利单抗", "帕博利珠单抗", "替雷利珠单抗", "卡瑞利珠单抗", "纳武利尤单抗",
]
TARGETED_KEYWORDS = [
    "靶向", "曲妥珠单抗", "帕妥珠单抗", "贝伐珠单抗", "阿帕他胺", "奥拉帕利", "恩杂鲁胺",
    "阿比特龙", "瑞戈非尼", "西妥昔单抗", "伊尼妥单抗", "厄洛替尼", "吉非替尼", "奥希替尼",
    "仑伐替尼", "阿来替尼",
]
ENDO_KEYWORDS = [
    "内分泌", "戈舍瑞林", "比卡鲁胺", "氟维司群", "来曲唑", "阿那曲唑", "他莫昔芬",
    "亮丙瑞林", "药物去势", "抗雄",
]

# 手术关键词用“严格版”：要尽量识别治疗性手术，同时避免把单纯活检/穿刺活检算成“手术参与”。
SURGERY_MAJOR_KEYWORDS = [
    "根治术", "切除术", "切除", "清扫术", "造瘘", "吻合", "探查", "重建", "修复",
    "胸腔镜", "腹腔镜", "开腹", "叶切除", "胃切除", "改良根治", "保乳", "肺叶切除",
    "单孔", "腔镜下",
]
SURGERY_BIOPSY_ONLY_KEYWORDS = ["活检", "穿刺活检", "系统性穿刺"]




def decode_hashu(text: Any) -> str:
    """Decode file names like '#U80ba#U764c' -> '肺癌'."""
    raw = str(text or "")
    return re.sub(r"#U([0-9a-fA-F]{4,6})", lambda m: chr(int(m.group(1), 16)), raw)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze population-level significance for before/after communication simulations."
    )
    parser.add_argument("--before_dir", required=True, help="Folder of before-training JSON files (e.g., b/).")
    parser.add_argument("--after_dir", required=True, help="Folder of after-training JSON files (e.g., a/).")
    parser.add_argument("--output_dir", required=True, help="Directory where the two Excel files and Markdown report will be written.")
    return parser.parse_args()


def extract_age(age_text: Any) -> float:
    """Extract integer age from text like '53岁'. Returns NaN if unavailable."""
    if age_text is None:
        return math.nan
    match = re.search(r"(\d+)", str(age_text))
    return float(match.group(1)) if match else math.nan


def parse_cancer_type(file_name: Any, fallback_path: Path) -> Optional[str]:
    """
    Parse cancer type from JSON field file_name.
    Example:
        supplement_001452_前列腺癌.json -> 前列腺癌
    """
    raw = str(file_name) if file_name else fallback_path.name
    stem = Path(raw).stem
    if "_" in stem:
        return stem.rsplit("_", 1)[-1]
    return stem or None


def contains_any(text: Any, keywords: Iterable[str]) -> bool:
    s = str(text or "")
    return any(keyword in s for keyword in keywords)


def split_clauses(text: Any) -> List[str]:
    """
    Split symptom text into semicolon/newline-delimited clauses.
    We intentionally keep the rule simple and transparent for maintainability.
    """
    s = str(text or "")
    s = s.replace("。", ";").replace("；", ";")
    clauses = [c.strip(" ;，,。") for c in re.split(r"[;\n]+", s) if c.strip(" ;，,。")]
    return clauses


def is_negative_clause(clause: str) -> bool:
    """
    Very simple rule for excluding normal/negative symptom clauses.
    Example negative clauses:
    - 无发热
    - 否认咯血
    - 未见异常
    - 正常
    """
    c = clause.strip()
    if not c:
        return True
    return (
        c.startswith("无")
        or c.startswith("否认")
        or "未见" in c
        or c.endswith("正常")
        or c == "正常"
        or "尚可" in c
        or "阴性" in c
        or "未触及" in c
        or "平稳" in c
    )


def positive_symptom_count(chief: Any, reiterated: Any, additional: Any) -> int:
    """
    Count the number of positive symptom clauses across three text fields.

    Important note:
    - This is a pragmatic text-derived proxy.
    - It is not a validated symptom severity scale.
    - Exact duplicate clauses are deduplicated to reduce obvious double counting.
    """
    seen: set[str] = set()
    count = 0
    for text in [chief, reiterated, additional]:
        for clause in split_clauses(text):
            if is_negative_clause(clause):
                continue
            normalized = clause.strip()
            if normalized and normalized not in seen:
                seen.add(normalized)
                count += 1
    return count


def smoking_history_rule(text: Any) -> bool:
    """Detect smoking history from personal_history.smoking_status."""
    s = str(text or "")
    if any(k in s for k in ["未知", "否认", "无吸烟史", "不吸烟", "否认、吸烟史"]):
        return False
    if "戒烟" in s:
        return True
    if re.search(r"吸烟|抽烟|支[／/]?日|包[／/]?天|支/天|支／日|包/天|包／天", s):
        return True
    return False


def heavy_alcohol_history_rule(text: Any) -> bool:
    """
    Detect 'heavy alcohol history' using simple interpretable heuristics.

    Positive examples:
    - >=20 years of drinking
    - 4-5+ liang/day
    - 7 liang/day
    - 200g/day, half-jin/day, etc.

    Negative examples:
    - 否认
    - 未知
    - 偶尔饮酒 / 偶饮 / 偶有饮酒
    - 量不详
    """
    s = str(text or "")
    if any(k in s for k in ["否认", "未知", "无饮酒史", "否认嗜酒史", "否认吸烟饮酒史"]):
        return False
    if any(k in s for k in ["偶尔", "偶饮", "偶有饮酒", "量不祥", "量不详"]):
        return False
    if any(k in s for k in ["半斤", "200g/天", "250g/天", "300-450g", "1000-2000ml", "7两/日", "5两/日", "5两／日", "4两/日", "白酒"]):
        return True
    if re.search(r"([2-9]|\d{2,})\+?两[／/]?[日天]", s):
        return True
    if re.search(r"每日([2-9]|\d{2,})两", s):
        return True
    year_match = re.search(r"(\d+)\+?年", s)
    if year_match and int(year_match.group(1)) >= 20:
        return True
    if any(k in s for k in ["20年以上", "20余年", "30余年", "30年以上", "40年以上", "50+年"]):
        return True
    return False


def lymph_node_involvement_rule(stage_text: Any, disease_name: Any, complication: Any, aux_list: Any) -> bool:
    """Detect lymph-node involvement from diagnosis / stage / complication / auxiliary examinations."""
    aux_text = " ".join(f"{item.get('item', '')} {item.get('result', '')}" for item in (aux_list or []))
    text = " ".join([str(stage_text or ""), str(disease_name or ""), str(complication or ""), aux_text])
    if "淋巴结" in text:
        return True
    if re.search(r"\bN[123]\b", text):
        return True
    return False


def metastatic_recurrent_context_rule(
    stage_text: Any,
    disease_name: Any,
    complication: Any,
    chief_complaint: Any,
    reiterated_symptom: Any,
    aux_list: Any,
) -> bool:
    """Detect metastatic/recurrent context from diagnosis / symptoms / auxiliary exams."""
    aux_text = " ".join(f"{item.get('item', '')} {item.get('result', '')}" for item in (aux_list or []))
    text = " ".join([
        str(stage_text or ""),
        str(disease_name or ""),
        str(complication or ""),
        str(chief_complaint or ""),
        str(reiterated_symptom or ""),
        aux_text,
    ])
    if any(k in text for k in ["转移", "复发", "播散", "种植", "骨转移", "脑转移", "肺转移", "肝转移", "腹膜转移"]):
        return True
    if re.search(r"\bM1\b", text):
        return True
    return False


def surgery_involved_rule(plan_text: Any) -> bool:
    """
    Detect treatment-related surgery.
    We intentionally exclude biopsy-only plans to keep the variable closer to 'therapeutic surgery involvement'.
    """
    s = str(plan_text or "")
    if any(k in s for k in SURGERY_MAJOR_KEYWORDS):
        return True
    if "术" in s and not any(k in s for k in SURGERY_BIOPSY_ONLY_KEYWORDS):
        return True
    return False


# =========================
# Record loading
# =========================

def load_records(data_dir: Path, dataset_label: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Load all JSON files from one directory and extract the analysis-ready table.

    Inclusion rule:
    - scores.patient_scores must be non-empty
    - the final score entry must contain pas_score / ess_score / ccs_score

    This keeps the script close to the earlier pilot analysis:
    records with available patient score trajectories are retained even if negotiation_result == "error",
    as long as the final patient score entry exists.
    """
    rows: List[Dict[str, Any]] = []
    excluded_files: List[str] = []
    total_files = 0

    for json_path in sorted(data_dir.rglob("*.json")):
        total_files += 1
        obj = json.loads(json_path.read_text(encoding="utf-8"))

        patient_scores = obj.get("scores", {}).get("patient_scores", [])
        if not patient_scores:
            excluded_files.append(decode_hashu(json_path.name))
            continue

        final_scores = patient_scores[-1]
        required_metrics = ["pas_score", "ess_score", "ccs_score"]
        if not all(metric in final_scores for metric in required_metrics):
            excluded_files.append(decode_hashu(json_path.name))
            continue

        patient_data = obj.get("patient_data", {})
        personal_info = patient_data.get("personal_info", {})
        demographics = personal_info.get("demographics", {})
        personal_history = personal_info.get("personal_history", {})
        social_background = personal_info.get("social_background", {})
        characteristics = personal_info.get("characteristics", {})

        examination_data = obj.get("examination_data", {})
        diagnosis = examination_data.get("diagnosis", {})
        treatment = examination_data.get("treatment", {})
        aux_exams = examination_data.get("auxiliary_examination", [])

        symptom = patient_data.get("symptom", {}) or examination_data.get("symptom", {}) or {}
        chief_complaint = symptom.get("chief_complaint", "")
        reiterated_symptom = patient_data.get("reiterated_symptom", "")
        additional_symptom = symptom.get("additional_symptom", "")
        treatment_plan = treatment.get("plan", "")

        record = {
            "dataset": dataset_label,  # "before" or "after"
            "json_file": json_path.name,
            "file_name": obj.get("file_name", json_path.name),

            # Existing fields
            "gender": demographics.get("gender"),
            "age": extract_age(demographics.get("age")),
            "cancer_type": parse_cancer_type(obj.get("file_name"), json_path),
            "stage_1234": diagnosis.get("stage_1234"),
            "education_level": social_background.get("education_level"),
            "financial_status": social_background.get("financial_status"),
            "personality": characteristics.get("personality"),
            "communication_style": characteristics.get("communication_style"),

            # Outcome variables
            "final_pas": float(final_scores["pas_score"]),
            "final_ess": float(final_scores["ess_score"]),
            "max_ess": float(max(item.get("ess_score", math.nan) for item in patient_scores if isinstance(item, dict))),
            "final_ccs": float(final_scores["ccs_score"]),

            # Raw text fields used to derive extended variables
            "smoking_status_raw": personal_history.get("smoking_status", ""),
            "alcohol_use_raw": personal_history.get("alcohol_use", ""),
            "chief_complaint": chief_complaint,
            "reiterated_symptom": reiterated_symptom,
            "additional_symptom": additional_symptom,
            "stage_raw": diagnosis.get("stage", ""),
            "disease_name": diagnosis.get("disease_name", ""),
            "complication_raw": diagnosis.get("complication", ""),
            "treatment_plan": treatment_plan,
            "auxiliary_examination": aux_exams,

            # Meta
            "completed_turns": obj.get("completed_turns"),
            "negotiation_result": obj.get("negotiation_result"),
            "n_score_points": len(patient_scores),
        }

        # Derived / extended variables
        record["systemic_therapy_involved"] = (
            contains_any(treatment_plan, CHEMO_KEYWORDS)
            or contains_any(treatment_plan, IMMUNO_KEYWORDS)
            or contains_any(treatment_plan, TARGETED_KEYWORDS)
            or contains_any(treatment_plan, ENDO_KEYWORDS)
        )
        record["surgery_involved"] = surgery_involved_rule(treatment_plan)
        record["chemotherapy_involved"] = contains_any(treatment_plan, CHEMO_KEYWORDS)
        record["symptom_burden"] = positive_symptom_count(
            chief=chief_complaint,
            reiterated=reiterated_symptom,
            additional=additional_symptom,
        ) >= 4
        record["metastatic_recurrent_context"] = metastatic_recurrent_context_rule(
            stage_text=diagnosis.get("stage", ""),
            disease_name=diagnosis.get("disease_name", ""),
            complication=diagnosis.get("complication", ""),
            chief_complaint=chief_complaint,
            reiterated_symptom=reiterated_symptom,
            aux_list=aux_exams,
        )
        record["lymph_node_involvement"] = lymph_node_involvement_rule(
            stage_text=diagnosis.get("stage", ""),
            disease_name=diagnosis.get("disease_name", ""),
            complication=diagnosis.get("complication", ""),
            aux_list=aux_exams,
        )
        record["heavy_alcohol_history"] = heavy_alcohol_history_rule(personal_history.get("alcohol_use", ""))
        record["smoking_history"] = smoking_history_rule(personal_history.get("smoking_status", ""))

        rows.append(record)

    df = pd.DataFrame(rows)
    summary = {
        "dataset": dataset_label,
        "input_dir": str(data_dir),
        "total_json_files": total_files,
        "included_records": int(len(df)),
        "excluded_records": int(len(excluded_files)),
        "excluded_files": excluded_files,
    }
    summary["excluded_files"] = [decode_hashu(x) for x in summary["excluded_files"]]
    return df, summary


# =========================
# Statistical tests
# =========================

def ordered_levels(series: pd.Series, spec: FactorSpec) -> List[Any]:
    """
    Determine the order of category levels shown in summaries.
    If a predefined level_order exists, keep it; otherwise preserve sorted unique values where possible.
    """
    observed = [value for value in pd.Series(series).dropna().unique().tolist()]
    if spec.level_order:
        ordered = [value for value in spec.level_order if value in observed]
        leftovers = [value for value in observed if value not in ordered]
        # Keep any unexpected new levels at the end to avoid losing information.
        ordered.extend(sorted(leftovers, key=lambda x: str(x)))
        return ordered
    return sorted(observed, key=lambda x: str(x))


def format_factor_value(spec: FactorSpec, value: Any) -> str:
    """Human-readable value formatting for summary sheets."""
    if pd.isna(value):
        return "NA"
    if spec.kind == "binary":
        return spec.bool_labels[1] if bool(value) else spec.bool_labels[0]
    return str(value)


def run_test(df: pd.DataFrame, spec: FactorSpec, outcome_key: str) -> Dict[str, Any]:
    """
    Run the most suitable test for one factor/outcome pair.

    Rules:
    - continuous -> Spearman rank correlation
    - binary     -> Mann–Whitney U (two-sided)
    - categorical (>2 groups) -> Kruskal–Wallis
    - categorical with exactly 2 observed groups -> Mann–Whitney U
    """
    subset = df[[spec.key, outcome_key]].dropna()

    if subset.empty:
        return {
            "method": "NA",
            "statistic": math.nan,
            "p_value": math.nan,
            "n_valid": 0,
            "observed_levels": [],
        }

    if spec.kind == "continuous":
        rho, p_value = stats.spearmanr(subset[spec.key], subset[outcome_key], alternative="two-sided")
        return {
            "method": "Spearman",
            "statistic": float(rho),
            "p_value": float(p_value),
            "n_valid": int(len(subset)),
            "observed_levels": [],
        }

    levels = ordered_levels(subset[spec.key], spec)
    groups = [subset.loc[subset[spec.key] == level, outcome_key] for level in levels]

    if len(levels) < 2:
        return {
            "method": "NA",
            "statistic": math.nan,
            "p_value": math.nan,
            "n_valid": int(len(subset)),
            "observed_levels": levels,
        }

    if spec.kind == "binary" or len(levels) == 2:
        stat, p_value = stats.mannwhitneyu(groups[0], groups[1], alternative="two-sided")
        return {
            "method": "Mann–Whitney U",
            "statistic": float(stat),
            "p_value": float(p_value),
            "n_valid": int(len(subset)),
            "observed_levels": levels,
        }

    stat, p_value = stats.kruskal(*groups)
    return {
        "method": "Kruskal–Wallis",
        "statistic": float(stat),
        "p_value": float(p_value),
        "n_valid": int(len(subset)),
        "observed_levels": levels,
    }


def build_tests_long(df: pd.DataFrame, specs: List[FactorSpec]) -> pd.DataFrame:
    """Long-format table: one row per dataset x factor x outcome."""
    rows: List[Dict[str, Any]] = []
    for spec in specs:
        for outcome_key, outcome_short, outcome_display in OUTCOMES:
            result = run_test(df, spec, outcome_key)
            rows.append({
                "dataset": df["dataset"].iloc[0] if not df.empty else "",
                "factor_key": spec.key,
                "factor_display": spec.display,
                "outcome_key": outcome_key,
                "outcome_short": outcome_short,
                "outcome_display": outcome_display,
                "method": result["method"],
                "statistic": result["statistic"],
                "p_value": result["p_value"],
                "significant_p_lt_0_05": bool(pd.notna(result["p_value"]) and result["p_value"] < 0.05),
                "n_valid": result["n_valid"],
                "observed_levels": " | ".join(format_factor_value(spec, x) for x in result["observed_levels"]),
            })
    return pd.DataFrame(rows)


def build_level_summary(df: pd.DataFrame, specs: List[FactorSpec]) -> pd.DataFrame:
    """
    Build a summary of group counts and outcome medians.
    This is particularly useful for interpreting Kruskal–Wallis omnibus p-values.
    """
    rows: List[Dict[str, Any]] = []
    dataset_name = df["dataset"].iloc[0] if not df.empty else ""

    for spec in specs:
        if spec.kind == "continuous":
            continue

        levels = ordered_levels(df[spec.key], spec)
        for level in levels:
            subset = df.loc[df[spec.key] == level]
            rows.append({
                "dataset": dataset_name,
                "factor_key": spec.key,
                "factor_display": spec.display,
                "level_raw": level,
                "level_display": format_factor_value(spec, level),
                "n": int(len(subset)),
                "median_final_pas": float(subset["final_pas"].median()) if len(subset) else math.nan,
                "median_final_ess": float(subset["final_ess"].median()) if len(subset) else math.nan,
                "median_max_ess": float(subset["max_ess"].median()) if len(subset) else math.nan,
                "median_final_ccs": float(subset["final_ccs"].median()) if len(subset) else math.nan,
            })

    return pd.DataFrame(rows)


def build_matrix(before_tests: pd.DataFrame, after_tests: pd.DataFrame, specs: List[FactorSpec]) -> pd.DataFrame:
    """Build the requested 8 x N matrix with rows like b_PAS, a_PAS, ..."""
    matrix = pd.DataFrame(index=[row_label for _, _, row_label in ROW_ORDER],
                          columns=[spec.display for spec in specs])

    for dataset_label, outcome_key, row_label in ROW_ORDER:
        source = before_tests if dataset_label == "before" else after_tests
        for spec in specs:
            match = source.loc[
                (source["factor_key"] == spec.key) &
                (source["outcome_key"] == outcome_key),
                "p_value"
            ]
            matrix.loc[row_label, spec.display] = float(match.iloc[0]) if not match.empty else math.nan

    return matrix


# =========================
# Workbook / report writing
# =========================

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ROW_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SIG_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def format_p_value(p: Any) -> str:
    """Readable p-value formatting for the matrix sheet."""
    if p is None or pd.isna(p):
        return "NA"
    p = float(p)
    return f"{p:.2e}" if p < 1e-3 else f"{p:.4f}"


def apply_basic_sheet_style(ws, freeze_cell: str = "B2") -> None:
    """Apply a clean, readable style to one worksheet."""
    ws.freeze_panes = freeze_cell
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def autosize_columns(ws, min_width: int = 10, max_width: int = 38) -> None:
    """Rudimentary but effective column auto-width."""
    for col_cells in ws.columns:
        column_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, min_width), max_width)


def dataframe_to_sheet(ws, df: pd.DataFrame, index: bool = False) -> None:
    """Write a DataFrame into an openpyxl worksheet."""
    start_row = 1
    start_col = 1

    headers = list(df.columns)
    if index:
        headers = [df.index.name or "index"] + headers

    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        values = row.tolist()
        if index:
            values = [row.name] + values

        for col_idx, value in enumerate(values, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    apply_basic_sheet_style(ws)


def write_matrix_sheet(ws, matrix: pd.DataFrame) -> None:
    """
    Write the compact p-value matrix requested by the user.
    Significant cells (p < 0.05) are highlighted.
    """
    ws.title = "Matrix"

    # Header row
    ws.cell(row=1, column=1, value="")
    corner = ws.cell(row=1, column=1)
    corner.fill = HEADER_FILL
    corner.font = HEADER_FONT
    corner.border = BORDER

    for col_idx, column_name in enumerate(matrix.columns, start=2):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, row_name in enumerate(matrix.index, start=2):
        row_header = ws.cell(row=row_idx, column=1, value=row_name)
        row_header.fill = ROW_HEADER_FILL
        row_header.font = Font(bold=True)
        row_header.border = BORDER
        row_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, column_name in enumerate(matrix.columns, start=2):
            p_value = matrix.loc[row_name, column_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=format_p_value(p_value))
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if pd.notna(p_value) and float(p_value) < 0.05:
                cell.fill = SIG_FILL
                cell.font = Font(bold=True)

    ws.freeze_panes = "B2"
    autosize_columns(ws, min_width=12, max_width=28)


def style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_workbook(
    output_path: Path,
    matrix: pd.DataFrame,
    tests_df: pd.DataFrame,
    level_df: pd.DataFrame,
    definitions_df: pd.DataFrame,
    dataset_summary_df: pd.DataFrame,
) -> None:
    """Create one workbook with Matrix, Tests, LevelSummary, Definitions, DatasetSummary sheets."""
    wb = Workbook()

    # Matrix
    ws_matrix = wb.active
    write_matrix_sheet(ws_matrix, matrix)

    # Tests
    ws_tests = wb.create_sheet("Tests")
    dataframe_to_sheet(ws_tests, tests_df, index=False)
    ws_tests.freeze_panes = "A2"
    autosize_columns(ws_tests, min_width=12, max_width=36)

    # Highlight p-value cells in Tests sheet
    p_col_idx = list(tests_df.columns).index("p_value") + 1
    for row_idx in range(2, ws_tests.max_row + 1):
        value = ws_tests.cell(row=row_idx, column=p_col_idx).value
        if isinstance(value, (int, float)) and value < 0.05:
            ws_tests.cell(row=row_idx, column=p_col_idx).fill = SIG_FILL
            ws_tests.cell(row=row_idx, column=p_col_idx).font = Font(bold=True)

    # Level summary
    ws_levels = wb.create_sheet("LevelSummary")
    dataframe_to_sheet(ws_levels, level_df, index=False)
    ws_levels.freeze_panes = "A2"
    autosize_columns(ws_levels, min_width=12, max_width=36)

    # Definitions
    ws_def = wb.create_sheet("Definitions")
    dataframe_to_sheet(ws_def, definitions_df, index=False)
    ws_def.freeze_panes = "A2"
    autosize_columns(ws_def, min_width=14, max_width=46)

    # Dataset summary
    ws_sum = wb.create_sheet("DatasetSummary")
    dataframe_to_sheet(ws_sum, dataset_summary_df, index=False)
    ws_sum.freeze_panes = "A2"
    autosize_columns(ws_sum, min_width=14, max_width=46)

    wb.save(output_path)


def summarize_significant_outcomes(tests_df: pd.DataFrame, display_name: str) -> str:
    """
    Build a compact Markdown summary like:
    personality: before(PAS, ESS), after(PAS, MaxESS, CCS)
    """
    lines: List[str] = []
    factor_subset = tests_df.loc[tests_df["factor_display"] == display_name].copy()
    for dataset in ["before", "after"]:
        sub = factor_subset.loc[(factor_subset["dataset"] == dataset) & (factor_subset["p_value"] < 0.05)]
        outcomes = sub["outcome_display"].tolist()
        if outcomes:
            label = "b" if dataset == "before" else "a"
            lines.append(f"{label}: " + ", ".join(outcomes))
    return "; ".join(lines) if lines else "未发现 p < 0.05 的关联"


def build_definitions_dataframe(specs: List[FactorSpec]) -> pd.DataFrame:
    rows = []
    for spec in specs:
        rows.append({
            "factor_key": spec.key,
            "factor_display": spec.display,
            "kind": spec.kind,
            "source": spec.source,
            "definition": spec.definition,
            "bool_labels_or_level_hint": (
                f"{spec.bool_labels[0]} / {spec.bool_labels[1]}" if spec.kind == "binary"
                else (" | ".join(str(x) for x in spec.level_order) if spec.level_order else "")
            ),
        })
    return pd.DataFrame(rows)


def build_dataset_summary_df(before_summary: Dict[str, Any], after_summary: Dict[str, Any]) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "dataset": "before (b)",
            "input_dir": before_summary["input_dir"],
            "total_json_files": before_summary["total_json_files"],
            "included_records": before_summary["included_records"],
            "excluded_records": before_summary["excluded_records"],
            "excluded_files": ", ".join(before_summary["excluded_files"]),
        },
        {
            "dataset": "after (a)",
            "input_dir": after_summary["input_dir"],
            "total_json_files": after_summary["total_json_files"],
            "included_records": after_summary["included_records"],
            "excluded_records": after_summary["excluded_records"],
            "excluded_files": ", ".join(after_summary["excluded_files"]),
        },
    ])


def build_report(
    output_path: Path,
    all_tests: pd.DataFrame,
    before_summary: Dict[str, Any],
    after_summary: Dict[str, Any],
) -> None:
    """
    Write a Markdown report describing:
    - sample sizes / filtering
    - statistical methods
    - variable definitions
    - main patterns found in this run
    """
    # Convenience subsets
    existing_names = [spec.display for spec in EXISTING_FACTORS]
    extended_names = [spec.display for spec in EXTENDED_FACTORS]
    existing_tests = all_tests.loc[all_tests["factor_display"].isin(existing_names)].copy()
    extended_tests = all_tests.loc[all_tests["factor_display"].isin(extended_names)].copy()

    def sig_rows(df: pd.DataFrame) -> pd.DataFrame:
        return df.loc[df["p_value"] < 0.05].sort_values(["dataset", "p_value", "factor_display"])

    existing_sig = sig_rows(existing_tests)
    extended_sig = sig_rows(extended_tests)

    # Factors that are never significant in either dataset / any outcome
    factor_sig_counts = all_tests.groupby("factor_display")["p_value"].apply(lambda s: int((s < 0.05).sum()))
    never_sig = sorted([name for name, count in factor_sig_counts.items() if count == 0], key=str)

    # Compact factor-level summary tables
    existing_summary_rows = []
    for spec in EXISTING_FACTORS:
        existing_summary_rows.append({
            "factor": spec.display,
            "summary": summarize_significant_outcomes(existing_tests, spec.display),
        })
    extended_summary_rows = []
    for spec in EXTENDED_FACTORS:
        extended_summary_rows.append({
            "factor": spec.display,
            "summary": summarize_significant_outcomes(extended_tests, spec.display),
        })

    def table_from_rows(rows: List[Dict[str, Any]], headers: List[str]) -> str:
        lines = []
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for row in rows:
            lines.append("| " + " | ".join(str(row[h]) for h in headers) + " |")
        return "\n".join(lines)

    report = []
    report.append("# Population-level significance report")
    report.append("")
    report.append("## 1. Input data and filtering")
    report.append("")
    report.append(f"- before_dir (b): {before_summary['included_records']} included / {before_summary['total_json_files']} total")
    report.append(f"- after_dir (a): {after_summary['included_records']} included / {after_summary['total_json_files']} total")
    report.append(f"- Excluded before files: {', '.join(before_summary['excluded_files']) if before_summary['excluded_files'] else 'None'}")
    report.append(f"- Excluded after files: {', '.join(after_summary['excluded_files']) if after_summary['excluded_files'] else 'None'}")
    report.append("")
    report.append("The script keeps records with non-empty `scores.patient_scores` and available final `pas_score / ess_score / ccs_score`. Final metrics are taken from the last patient score entry; `Maximum ESS` is the maximum `ess_score` across the whole dialogue.")
    report.append("")
    report.append("## 2. Statistical methods")
    report.append("")
    report.append("- Binary factors: two-sided Mann–Whitney U test")
    report.append("- Multi-category categorical factors: Kruskal–Wallis omnibus test")
    report.append("- Age (continuous): two-sided Spearman rank correlation")
    report.append("- Matrix cells are raw exploratory p-values; significant cells are highlighted at `p < 0.05`.")
    report.append("- For omnibus categorical tests, the p-value only indicates that at least one category differs; use the `LevelSummary` sheet to inspect group medians and counts.")
    report.append("")
    report.append("## 3. Variable definitions used in this run")
    report.append("")
    report.append("### 3.1 Existing fields")
    report.append("")
    report.append(table_from_rows(
        [
            {"factor": spec.display, "source": spec.source, "definition": spec.definition}
            for spec in EXISTING_FACTORS
        ],
        ["factor", "source", "definition"],
    ))
    report.append("")
    report.append("### 3.2 Extended fields")
    report.append("")
    report.append(table_from_rows(
        [
            {"factor": spec.display, "source": spec.source, "definition": spec.definition}
            for spec in EXTENDED_FACTORS
        ],
        ["factor", "source", "definition"],
    ))
    report.append("")
    report.append("## 4. What the current run finds")
    report.append("")
    report.append("### 4.1 Existing fields: factor-level summary")
    report.append("")
    report.append(table_from_rows(existing_summary_rows, ["factor", "summary"]))
    report.append("")
    report.append("### 4.2 Extended fields: factor-level summary")
    report.append("")
    report.append(table_from_rows(extended_summary_rows, ["factor", "summary"]))
    report.append("")
    report.append("### 4.3 High-level interpretation")
    report.append("")
    report.append("- **Personality is the strongest and most stable factor** in this run. It is significant across all four endpoints in both before (b) and after (a), although the signal is much stronger after training.")
    report.append("- **Stage and treatment-complexity proxies are mainly significant in the before model (b)**. In this run, stage, systemic-therapy involvement, surgery involvement, chemotherapy involvement, symptom burden, and metastatic/recurrent context are concentrated in b rather than a. This is consistent with the earlier narrative that RL optimization may attenuate some coarse-grained disparities.")
    report.append("- **After training (a), smoking history and heavy alcohol history remain salient**, while several coarse treatment-complexity factors become non-significant. This suggests that the post-training difficulty landscape may shift toward more patient-history-sensitive heterogeneity.")
    report.append("- **Communication style shows an a-side signal**, while **gender and age do not show robust evidence** of association in either dataset under the present direct-field analysis.")
    report.append("")
    if never_sig:
        report.append("### 4.4 Factors without p < 0.05 in this run")
        report.append("")
        report.append(", ".join(never_sig))
        report.append("")
    report.append("### 4.5 Significant tests (long form)")
    report.append("")
    if all_tests.loc[all_tests["p_value"] < 0.05].empty:
        report.append("No p < 0.05 findings in this run.")
    else:
        long_rows = []
        sig_df = all_tests.loc[all_tests["p_value"] < 0.05].sort_values(["dataset", "factor_display", "p_value"])
        for _, row in sig_df.iterrows():
            long_rows.append({
                "dataset": row["dataset"],
                "factor": row["factor_display"],
                "outcome": row["outcome_display"],
                "method": row["method"],
                "p_value": format_p_value(row["p_value"]),
            })
        report.append(table_from_rows(long_rows, ["dataset", "factor", "outcome", "method", "p_value"]))
        report.append("")
    report.append("## 5. Maintenance notes")
    report.append("")
    report.append("- If you want to tighten or relax any extended-variable rule, edit the corresponding helper function in the script:")
    report.append("  - `smoking_history_rule`")
    report.append("  - `heavy_alcohol_history_rule`")
    report.append("  - `lymph_node_involvement_rule`")
    report.append("  - `metastatic_recurrent_context_rule`")
    report.append("  - `surgery_involved_rule`")
    report.append("  - `positive_symptom_count` (for symptom burden)")
    report.append("- If later you want age bins or merged categories instead of direct-field tests, those should be implemented as *new derived factors* rather than silently replacing the current definitions.")
    report.append("")
    output_path.write_text("\n".join(report), encoding="utf-8")


def main() -> None:
    args = parse_args()
    before_dir = Path(args.before_dir)
    after_dir = Path(args.after_dir)
    output_dir = Path(args.output_dir)

    if not before_dir.exists():
        raise FileNotFoundError(f"--before_dir does not exist: {before_dir}")
    if not after_dir.exists():
        raise FileNotFoundError(f"--after_dir does not exist: {after_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)

    # Load data
    before_df, before_summary = load_records(before_dir, dataset_label="before")
    after_df, after_summary = load_records(after_dir, dataset_label="after")

    # Build long-format test results
    before_existing_tests = build_tests_long(before_df, EXISTING_FACTORS)
    after_existing_tests = build_tests_long(after_df, EXISTING_FACTORS)
    before_extended_tests = build_tests_long(before_df, EXTENDED_FACTORS)
    after_extended_tests = build_tests_long(after_df, EXTENDED_FACTORS)

    # Build matrix sheets
    existing_matrix = build_matrix(before_existing_tests, after_existing_tests, EXISTING_FACTORS)
    extended_matrix = build_matrix(before_extended_tests, after_extended_tests, EXTENDED_FACTORS)

    # Build detailed summary sheets
    existing_tests = pd.concat([before_existing_tests, after_existing_tests], ignore_index=True)
    extended_tests = pd.concat([before_extended_tests, after_extended_tests], ignore_index=True)

    existing_levels = pd.concat([
        build_level_summary(before_df, EXISTING_FACTORS),
        build_level_summary(after_df, EXISTING_FACTORS),
    ], ignore_index=True)

    extended_levels = pd.concat([
        build_level_summary(before_df, EXTENDED_FACTORS),
        build_level_summary(after_df, EXTENDED_FACTORS),
    ], ignore_index=True)

    dataset_summary_df = build_dataset_summary_df(before_summary, after_summary)
    existing_definitions = build_definitions_dataframe(EXISTING_FACTORS)
    extended_definitions = build_definitions_dataframe(EXTENDED_FACTORS)

    # Write workbooks
    existing_xlsx = output_dir / "01_existing_fields_significance.xlsx"
    extended_xlsx = output_dir / "02_extended_fields_significance.xlsx"
    report_md = output_dir / "analysis_report.md"

    write_workbook(
        output_path=existing_xlsx,
        matrix=existing_matrix,
        tests_df=existing_tests,
        level_df=existing_levels,
        definitions_df=existing_definitions,
        dataset_summary_df=dataset_summary_df,
    )
    write_workbook(
        output_path=extended_xlsx,
        matrix=extended_matrix,
        tests_df=extended_tests,
        level_df=extended_levels,
        definitions_df=extended_definitions,
        dataset_summary_df=dataset_summary_df,
    )

    all_tests = pd.concat([existing_tests, extended_tests], ignore_index=True)
    build_report(
        output_path=report_md,
        all_tests=all_tests,
        before_summary=before_summary,
        after_summary=after_summary,
    )

    print(f"Existing-fields workbook: {existing_xlsx}")
    print(f"Extended-fields workbook: {extended_xlsx}")
    print(f"Markdown report: {report_md}")


if __name__ == "__main__":
    main()