import json
import random
import re
from pathlib import Path

from openpyxl import load_workbook
from tqdm import tqdm

NAMES = ["胃癌", "结直肠癌", "前列腺癌"]
SUBDIRS = ["experiment", "statistics", "test", "train"]

DIAG_PATH = Path("../../data/diagnosis")

def get_dir(name: str) -> Path:
    return Path("../../data/raw") / f"{name}_preprocessed/success"

def get_excel(name: str) -> Path:
    return Path("../../data/raw/xlsx") / f"{name}病历_脱敏.xlsx"

def find_gender_and_name(disease_name: str, id: int) -> tuple[str, str]:
    print(f"Finding gender and age for {disease_name} with id {id}")

    wb = load_workbook(get_excel(disease_name))
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    # rows 能按照 rows['patientid'] 为第一关键字，rows['病历时间'] 作为第二关键字排序吗？
    sorted_rows = sorted(rows, key=lambda row: (dict(zip(headers, row, strict=False)).get("patientid", ""), dict(zip(headers, row, strict=False)).get("病历时间", "")))

    # 只选 row['patientid'] == id 的行
    selected_rows = [row for row in sorted_rows if dict(zip(headers, row, strict=False)).get("patientid", "") == id]
    for row in selected_rows:
        content = dict(zip(headers, row, strict=False)).get("文书内容", "")
        if content is None:
            # print(f"文书内容 is None for {disease_name} with id {id}")
            continue
        def _f(text: str):
            pattern = r"患者(?P<name>[^，,]+)[，,]\s*(?P<gender>男|女)[，,]\s*(?P<age>\d+)岁"
            
            match = re.search(pattern, text)
            if not match:
                return None, None, None
            
            name = match.group("name")
            gender = match.group("gender")
            age = int(match.group("age"))
            
            return name, gender, age
        name, gender, age = _f(content)
        if gender not in ["男", "女"] or age is None:
            # print(f"Could not extract gender and age from content for {disease_name} with id {id}")
            continue
        else:
            # print(f"Name: {name}, Gender: {gender}, Age: {age} for id {id}")
            return gender, f"{age}岁"

    if disease_name == "前列腺癌":
        # print("Another Chance")
        for row in selected_rows:
            content = dict(zip(headers, row, strict=False)).get("文书内容", "")
            # print(content)
            if content is None:
                # print(f"文书内容 is None for {disease_name} with id {id}")
                continue
            # 匹配 xxxx年龄：60岁yyyyy
            age_match = re.search(r"年龄：(\d+)岁", content)
            if age_match:
                age = int(age_match.group(1))
                # print(f"Extracted age: {age} for id {id}")
                return "男", f"{age}岁"
    else:
        # print("Another Chance")
        gender, age = None, None
        for row in selected_rows:
            content = dict(zip(headers, row, strict=False)).get("文书内容", "")
            if content is None:
                # print(f"文书内容 is None for {disease_name} with id {id}")
                continue
            # 匹配 xxxx年龄：60岁yyyyy
            age_match = re.search(r"年龄：(\d+)岁", content)
            if age_match:
                age = int(age_match.group(1))
                break
        if age is not None:
            # print(f"Extracted gender: {gender}, age: {age} for id {id}")
            return random.choice(["男", "女"]), f"{age}岁"
    print(f"Could not find a valid row for {disease_name} with id {id}")
    return "", ""

if __name__ == '__main__':
    for name in NAMES:
        dir = get_dir(name)
        for file in tqdm(list(dir.glob("**/*.json")), desc=f"Processing {name}"):
            data = json.load(file.open())
            flag = True
            if data["personal_info"]["demographics"]["gender"] not in ["男", "女"]:
                # print(f"Invalid gender value: {file}")
                flag = False
            if not data["personal_info"]["demographics"]["age"].split("岁")[0].isdigit():
                # print(f"Missing age value: {file}")
                flag = False
            if not flag:
                # print(f"Trying to find gender and age for {file}")
                gender, age = find_gender_and_name(name, int(file.stem))
                data["personal_info"]["demographics"]["gender"] = gender
                data["personal_info"]["demographics"]["age"] = age

                # 开始覆盖原有错误文件
                # 首先是 raw 里的文件
                with open(file, "w") as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)
                # 其次是 diagnosis 里的文件
                for subdir in SUBDIRS:
                    diag_path = DIAG_PATH / name / subdir
                    for diag_file in diag_path.glob("**/*.json"):
                        diag_data = json.load(diag_file.open())
                        # 如果里面的症状、辅助检查、诊断和治疗都和这个文件里的一样，就说明这个文件就是我们要找的那个文件
                        if diag_data["symptom"] == data["symptom"] \
                            and diag_data["auxiliary_examination"] == data["auxiliary_examination"] \
                                and diag_data["diagnosis"] == data["diagnosis"] \
                                    and diag_data["treatment"] == data["treatment"]:
                            diag_data["personal_info"]["demographics"]["gender"] = gender
                            diag_data["personal_info"]["demographics"]["age"] = age
                            with open(diag_file, "w") as f:
                                json.dump(diag_data, f, ensure_ascii=False, indent=4)
